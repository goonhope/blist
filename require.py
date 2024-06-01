# -*- coding: utf-8 -*-
"""
@Filename	:	require.py
@Created 	:	2024/03/28 15:17:52
@Updated	:	2024/03/28 15:17:52
@Author   	:	goonhope@gmail.com; Teddy; Zhuhai
@Function	:	required function ground
@Process  	:	read -> found -> extract -> done
@WitNote	:	by re
@Reference :  Personal project
"""
from faker import Faker
from functools import wraps
from bs4 import BeautifulSoup
import os, requests, cchardet, os.path as op, random, time


CNUM = {'广州': (200, '440100'), '韶关': (751, '440200'), '深圳': (755, '440300'),
                '珠海': (756, '440400'), '汕头': (754, '440500'), '佛山': (757, '440600'),
                '江门': (750, '440700'), '湛江': (759, '440800'), '茂名': (668, '440900'),
                '肇庆': (758, '441200'), '惠州': (752, '441300'), '梅州': (753, '441400'),
                '汕尾': (660, '441500'), '河源': (762, '441600'), '阳江': (662, '441700'),
                '清远': (763, '441800'), '东莞': (769, '441900'), '中山': (760, '442000'),
                '潮州': (768, '445100'), '揭阳': (663, '445200'), '云浮': (766, '445300')}


def fetch(url="", hdrs=None, data=None, proxy=None, json=False,g=True,raw=False, tout=5):
    """统一get post 默认get"""
    url_headers = google_hder(url.split("/")[2])
    if hdrs and isinstance(hdrs, dict): url_headers.update(hdrs)
    proxy = random.choice(proxy) if isinstance(proxy, list) else proxy
    way = requests.get if g else requests.post
    kw = dict(headers=url_headers, timeout=tout, proxies=proxy, verify=False)
    kw.update({"params" if g else "data": data})
    data = way(url, **kw)
    if data.status_code == 200:
        data.encoding = cchardet.detect(data.content)['encoding']  # 网页编码utf8 or GB18030
        return data.text if raw else (data.json() if json else BeautifulSoup(data.text, 'lxml'))
    else: return print("@fetch check !")


def excel(dir="", na="", t=True, r=True):
    """写入excel装饰器：xlsx or xls"""
    def ex_cel(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            start = time.time()
            info = func(*args, **kwargs)
            _time_ = time.strftime("_%Y%m%d_%H%M%S") if t else ""
            titles = [x for x in list(kwargs.values()) + list(args) if x and isinstance(x, str) and len(x) < 6]
            farg = (f"{titles[0]}_" if titles else "") + na or "file"
            out = _time_.join(op.splitext(farg)) if ".xlsx" in na.lower() else f'{farg}{_time_}.xlsx'
            file_out = op.join(dir or os.path.dirname(__file__), out)
            excel_in(info, file_out, rc=r,creator="goonhope@gmail.com",title="备案信息")
            print(f"@{func.__name__}:\t[Time:{time.time() - start : 0.1f}s]")
            return info
        return wrapper
    return ex_cel


def excel_in(info, file_out, rc=True,bg="ffffff",**kwargs):
    """xlsx 写入—背景颜色可选"""
    from openpyxl import Workbook,styles
    wb = Workbook()
    cfill = lambda x: styles.PatternFill("solid", fgColor=x)
    table = wb.create_sheet('list', 0)
    for row, rows in enumerate(info, 1):  # 行 y-->Row
        if isinstance(rows, str): rows = rows.split("\t")  # 避免字符串被分解
        for col, data in enumerate(rows, 1):  # 列 x--> Column
            x = row, col
            tcell = table.cell(*(x if rc else x[::-1]), value=data)
            tcell.font = styles.Font("微软雅黑")
            if bg.strip("f"): tcell.fill = cfill(bg)
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    meta(wb.properties, **kwargs)
    wb.save(file_out)


def meta(prop, **kwargs):
    """元信息修改"""
    for x, y in kwargs.items():
        if x in prop.__class__.__dict__ and prop.__getattribute__(x) != y:
            prop.__setattr__(x, y)


def google_hder(host=None, o=True):
    """'google search url headers"""
    google_hders = {
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN, zh;q=0.9',
        'Connection': 'keep-alive',
        'Referer': f'https://www.{"google.com.hk" if o else "qq.com"}',
        'Upgrade-insecure-requests': '1',
        'User-Agent': Faker("zh_CN").chrome()} # fr..user_agent()
    if host and isinstance(host,(str,dict)):
        google_hders.update(host if isinstance(host,dict) else dict(Host=host))
    return google_hders


def json_to_list(url, js=True, data=None):
    """url json to xls or soup"""
    url_headers = google_hder({'Host': url.split("/")[2]},False)
    if data and isinstance(data, dict): url_headers.update(data)
    dt = requests.get(url, headers=url_headers)
    if dt.status_code == 200:
        if js: return dt.json()
        else:
            dt.encoding = cchardet.detect(dt.content)['encoding']  # 确定网页编码 utf8 or GB18030
            return BeautifulSoup(dt.text, 'lxml')
    else: return False


def open_txt(keys="", file="delaw.log", llst=True, encoding="utf-8"):
    """读取txt 写文件，llist 1 读写二维list文件，0直接读写"""
    file = op.join(os.path.dirname(__file__), file) if not os.path.isabs(file) else file
    if keys:
        with open(file, "w", encoding=encoding) as f:
            f.write(list2d(keys) if llst else keys)
    else:
        if not op.exists(file): return []
        else:
            with open(file, "r", encoding=encoding) as f:
                info = f.read()
                return list2d(info) if llst else info


def list2d(llist, stp=r"'\" ", sep="\t"):
    """二维数列与字符串转换"""
    if isinstance(llist, list):
        return "\n".join(sep.join(str(x).strip(stp) for x in key) for key in llist)
    elif isinstance(llist, str):
        return [x.strip(stp).split(sep) for x in llist.strip().split("\n") if x.strip()] if llist else []
    else:
        return False


def post_(url, hdrs, data=None, json=True, proxy=None):
    """post"""
    proxy = random.choice(proxy) if isinstance(proxy, list) else proxy
    url_data = requests.post(url, data=data, headers=hdrs, timeout=3, proxies=proxy)
    if url_data.status_code == 200:
        url_data.encoding = cchardet.detect(url_data.content)['encoding']  # 网页编码utf8 or GB18030
        return url_data.json() if json else BeautifulSoup(url_data.text, 'lxml')
    else: return print("@post_ check !") or False


def post_while(url, hdrs, data, json=True, proxy=None, max=5):
    """post_while"""
    times, ifs = 0, False
    proxy = random.choice(proxy) if isinstance(proxy, list) else proxy
    while not ifs and times < max:
        try:
            url_data = requests.post(url, data=data, headers=hdrs, timeout=3, proxies=proxy)
            ifs = url_data.status_code == 200
            if ifs and times == max:
                url_encoding = cchardet.detect(url_data.content)['encoding']  # 确定网页原编码
                url_data.encoding = url_encoding  # ' utf8' or "GB18030"
                data_content = url_data.json() if json else BeautifulSoup(url_data.text, 'lxml')
                return data_content
        except:
            times += 1; print(f"@post_while, Try: {str(times)} times !")
            time.sleep(random.uniform(1, 3.14))


def time_from(t=0, fmt="%Y%m%d %H:%M:%S", local=True):
    """时间相互转换，时间戳与字符串"""
    if isinstance(t, (int, float)):
        flocal = time.gmtime if local else time.localtime
        delta = 0 if local else 8 * 60 ** 2
        return time.strftime(fmt, flocal(t + delta)) if t else time.strftime(fmt)
    elif isinstance(t, str):
        timeArray = time.strptime(t, fmt)
        return int(time.mktime(timeArray))
    else: return print("@t: int, float, str !") or False


def tsleep(max=0.,min=0.5):
    """固定随机sleep时间"""
    time.sleep(random.uniform(min, max) if max > min else min)


if __name__ == '__main__':
    print(time_from())
